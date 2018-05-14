import sys
from openpyxl import load_workbook
from utilities.CommonUtil import get_project_path
from robot.api.logger import *


class ExcelOperations:
    """
    Class Name -  ExcelOperations
    Description - This class contains some common methods used in framework
    development
    Parameters - None
    Return - None
    Author -  Dhananjay Joshi
    Modification date - 17-Apr-2018
    """
    @staticmethod
    def get_cell_data_from_excel(wb_name, sheet_name, row_num, col_num):
        """
        Function Name -  get_cell_data_from_excel
        Description - gets cell data from excel
        Parameters - wb_name, sheet_name, row_num, col_num
        Return - None
        Author -  Dhananjay Joshi
        Modification date - 24-Apr-2018
        """

        try:
            wb_name = get_project_path() + "\\testdata\\" + wb_name
            wb = load_workbook(wb_name)
            sheet = wb[sheet_name]
            val = sheet.cell(row=int(row_num), column=int(
                col_num)).internal_value
            info("Data read success: value is %s:: " % val, True, True)
            return val
        except Exception as e:
            tb = sys.exc_info()[2]
            error("Exception in reading excel data in "
                  "def get_cell_data_from_excel as  [%s]::Line Number[%s]" %
                  (e.with_traceback(tb), sys.exc_info()[2].tb_lineno), True)
            return "NA"