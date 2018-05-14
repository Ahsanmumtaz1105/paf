import sys
from openpyxl import load_workbook
from libs.utils.CommonUtil import get_project_path
from robot.api.logger import *


class ExcelOperations:
    """
    """
    @staticmethod
    def get_cell_data_from_excel(wb_name, sheet_name, row_num, col_num):
        """
        """

        try:
            wb_name = get_project_path() + "\\test_data\\" + wb_name
            wb = load_workbook(wb_name)
            sheet = wb[sheet_name]
            val = sheet.cell(row=int(row_num), column=int(
                col_num)).internal_value
            info("Data read success: value is %s:: " % val, True, True)
            return val
        except Exception as e:
            tb = sys.exc_info()[2]
            error("Exception in reading excel data in "
                  "def get_cell_data_from_excel as  [%s]::Line Number[%s]" %
                  (e.with_traceback(tb), sys.exc_info()[2].tb_lineno), True)
            return "NA"
